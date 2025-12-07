import json
import csv
from openpyxl import load_workbook

def read_in(file_path) :
    """
    Read in json files.
    """
    with open(file_path, "r", encoding="utf-8") as fin :
        data = json.load(fin)
    return data


def read_in_lines(file_path) :
    """
    Read in dicts in lines.
    """
    data = []
    with open(file_path, "r", encoding="utf-8") as fin :
        for idx, line in enumerate(fin) :
            try :
                data.append(json.loads(line))
            except :
                continue
    return data


def read_csv(file_path) :
    """
    Read in csv files.
    """
    data = []
    with open(file_path, "r", encoding="utf-8") as file :
        lines = file.readlines()
    
    if not lines :
        return []
    
    header_line = lines[0].strip()
    headers = header_line.split("\t")

    for i in range(1, len(lines)) :
        line = line[i].strip()
        if line :
            values = line.split("\t")
            row_dict = {}
            for j in range(min(len(headers), len(values))) :
                row_dict[headers[j]] = values[j]
            data.append(row_dict)
    return data

def read_jsonl_file(file_path) :
    with open(file_path, "r", encoding='utf-8') as fin :
        lines = fin.readlines()
    return lines

def read_plain_csv(file_path) :
    """
    Read a csv fiel and convert it into a markdown table format.

    Args:
        file_path (str): Path to the csv file

    Returns:
        str: Markdown formatted table
    """

    try :
        with open(file_path, "r", encoding="utf-8") as f :
            reader = csv.reader(f)
            rows = list(reader)

            if not rows :
                return "Empty CSV file"
            
            headers = rows[0]
            data_rows = rows[1:]

            # Add header row
            markdown_table = []
            markdown_table.append("| " + " | ".join(headers) + " |")

            # Add separator row
            markdown_table.append("| " + " | ".join(["---" for _ in headers]) + " |")

            # Add data row
            for row in data_rows :
                markdown_table.append("| " + " | ".join(row) + " |")

            return "\n".join(markdown_table)

    except FileNotFoundError :
        return f"Error: File {file_path} not Found"
    except Exception as e :
        return f"Error reading CSV file: {str(e)}"


def read_excel_to_markdown(file_path):
    # 1. data_only=True 确保读取的是值而不是公式
    workbook = load_workbook(file_path, data_only=True)

    content = []
    file_name = file_path.split("/")[-1]
    table_name = file_name.replace(".xlsx", "")

    content.append(f"Table name: {table_name}")

    # 获取所有 sheet 名称
    sheet_names = workbook.sheetnames
    # 判断是否需要显示 Sheet 标题（多于1个才显示）
    show_sheet_title = len(sheet_names) > 1

    for sheet_name in sheet_names:
        work_sheet = workbook[sheet_name]

        # 仅在多 Sheet 模式下添加 Sheet 标题
        if show_sheet_title:
            content.append(f"\n### Sheet: {sheet_name}")

        rows = list(work_sheet.iter_rows(values_only=True))
        if not rows:
            content.append("*Empty Sheet*")
            continue

        for i, row in enumerate(rows):
            # 2. 修复类型错误和None值
            clean_row = [str(cell) if cell is not None else "" for cell in row]
            # 处理换行符
            clean_row = [cell.replace("\n", "<br>") for cell in clean_row]

            content.append("| " + " | ".join(clean_row) + " |")

            # 3. 添加表头分隔符
            if i == 0:
                content.append("| " + " | ".join(["---"] * len(clean_row)) + " |")

    return "\n".join(content)


if __name__ == '__main__' :
    print(read_excel_to_markdown("../data/dev_excel/1st_New_Zealand_Parliament_0.xlsx"))